"""
tracker.py — the local Excel tracker (OpenPyXL). Single source of truth.

Schema (sheet 'Applications') matches blueprint §8. The ONLY way an "applied"
row is written is via append_application(), which the UI calls when you click
the ✅ tick — the human confirmation gate.
"""
import os
from datetime import date, timedelta

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

COLUMNS = [
    "app_id", "date_added", "company", "job_title", "location", "job_link",
    "source", "ats_score", "h1b_sponsor", "resume_file", "applied",
    "applied_date", "status", "follow_up_date", "notes",
]
SHEET = "Applications"
HEADER_FILL = PatternFill("solid", fgColor="11243F")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def ensure_workbook(path: str):
    if os.path.exists(path):
        return
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.append(COLUMNS)
    for i, _ in enumerate(COLUMNS, start=1):
        c = ws.cell(row=1, column=i)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    _autosize(ws)
    wb.save(path)


def _autosize(ws):
    widths = {1: 16, 2: 12, 3: 16, 4: 28, 5: 18, 6: 30, 7: 12, 8: 10,
              9: 12, 10: 30, 11: 9, 12: 12, 13: 12, 14: 14, 15: 30}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def next_app_id(path: str) -> str:
    today = date.today().strftime("%Y%m%d")
    n = 1
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb[SHEET]
        todays = [r[0].value for r in ws.iter_rows(min_row=2)
                  if r[0].value and str(r[0].value).startswith(f"APP-{today}")]
        n = len(todays) + 1
    return f"APP-{today}-{n:03d}"


def append_application(path: str, record: dict) -> str:
    """Append one application row. Fills defaults and stamps dates.
    Returns the app_id."""
    ensure_workbook(path)
    wb = load_workbook(path)
    ws = wb[SHEET]

    app_id = record.get("app_id") or next_app_id(path)
    today = date.today()
    applied = bool(record.get("applied", True))
    applied_date = record.get("applied_date") or (today.isoformat() if applied else "")
    follow_up = record.get("follow_up_date")
    if not follow_up and applied:
        follow_up = (today + timedelta(days=7)).isoformat()

    row = {
        "app_id": app_id,
        "date_added": record.get("date_added") or today.isoformat(),
        "company": record.get("company", ""),
        "job_title": record.get("job_title", ""),
        "location": record.get("location", ""),
        "job_link": record.get("job_link", ""),
        "source": record.get("source", "Manual"),
        "ats_score": record.get("ats_score", ""),
        "h1b_sponsor": bool(record.get("h1b_sponsor", False)),
        "resume_file": record.get("resume_file", ""),
        "applied": applied,
        "applied_date": applied_date,
        "status": record.get("status", "Applied" if applied else "Saved"),
        "follow_up_date": follow_up or "",
        "notes": record.get("notes", ""),
    }
    ws.append([row[c] for c in COLUMNS])
    wb.save(path)
    return app_id


def find_duplicate(path: str, company: str, job_title: str, job_link: str = "") -> dict:
    """Return an existing row for the same job (same link, or same company+title),
    or None. Used to stop double-recording the same application."""
    c = (company or "").strip().lower()
    t = (job_title or "").strip().lower()
    link = (job_link or "").strip().lower()
    for r in read_all(path):
        rlink = str(r.get("job_link", "")).strip().lower()
        if link and rlink and link == rlink:
            return r
        if c and t and str(r.get("company", "")).strip().lower() == c \
                and str(r.get("job_title", "")).strip().lower() == t:
            return r
    return None


def read_all(path: str) -> list:
    if not os.path.exists(path):
        return []
    wb = load_workbook(path)
    ws = wb[SHEET]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r and any(v is not None for v in r):
            rows.append(dict(zip(COLUMNS, r)))
    return rows


def update_status(path: str, app_id: str, status: str = None, notes: str = None) -> bool:
    if not os.path.exists(path):
        return False
    wb = load_workbook(path)
    ws = wb[SHEET]
    idx = {c: i + 1 for i, c in enumerate(COLUMNS)}
    for r in ws.iter_rows(min_row=2):
        if r[0].value == app_id:
            if status is not None:
                r[idx["status"] - 1].value = status
            if notes is not None:
                r[idx["notes"] - 1].value = notes
            wb.save(path)
            return True
    return False


def follow_ups_due_from_rows(rows: list, on: date = None) -> list:
    on = on or date.today()
    due = []
    for row in rows:
        fu = row.get("follow_up_date")
        if not fu:
            continue
        try:
            d = date.fromisoformat(str(fu)[:10])
        except ValueError:
            continue
        if d <= on and str(row.get("status", "")).lower() not in ("offer", "rejected"):
            due.append(row)
    return due


def follow_ups_due(path: str, on: date = None) -> list:
    return follow_ups_due_from_rows(read_all(path), on)


def summary_counts_from_rows(rows: list) -> dict:
    counts = {}
    for r in rows:
        s = r.get("status", "Unknown") or "Unknown"
        counts[s] = counts.get(s, 0) + 1
    counts["_total"] = len(rows)
    return counts


def summary_counts(path: str) -> dict:
    return summary_counts_from_rows(read_all(path))
